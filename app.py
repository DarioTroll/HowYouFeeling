from flask import Flask, render_template, request, redirect, jsonify, session, url_for, send_file
from dotenv import load_dotenv
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook
from sleepAmountUpdater import calcola_ore_sonno
import openpyxl
import os
import mysql.connector
import hashlib
import pandas as pd


load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", os.urandom(24))

db_host = os.getenv("DB_HOST")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_name = os.getenv("DB_NAME")

# Connessione al database
conn = mysql.connector.connect(
    host=db_host,
    user=db_user,
    password=db_password,
    database=db_name
)
cursor = conn.cursor(dictionary=True)

# Try revert that hash...
users = {
    "dariotrollo": "61b0e03adc0eb4bba073fbc584e5ac23a195a49676fb21976acb055f8d4ecc45"
}

def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

# --- LOGIN ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = "dariotrollo"
        password = request.form['password']
        if username in users and users[username] == hash_password(password):
            session['username'] = username
            return redirect(url_for('home'))
        else:
            return render_template("login.html", error="Password errata")
    return render_template('login.html')

# --- LOGOUT ---
@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

# --- DECORATORE PER PROTEGGERE LE PAGINE ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- HOME ---
@app.route("/")
@login_required
def home():
    return render_template("home.html")


# --- INSERIMENTI ---
@app.route('/inserisci_dolore', methods=['GET', 'POST'])
@login_required
def inserisci_dolore():
    if request.method == 'POST':
        dati = {
            "data": request.form['data'],
            "ora": request.form['ora'],
            "viso": request.form.get('viso', 0),
            "testa_fronte": request.form.get('testa_fronte', 0),
            "testa_tempie_dx": request.form.get('testa_tempie_dx', 0),
            "testa_tempie_sx": request.form.get('testa_tempie_sx', 0),
            "testa_sommità": request.form.get('testa_sommità', 0),
            "testa_occipite": request.form.get('testa_occipite', 0),
            "nuca": request.form.get('nuca', 0),
            "mascella_dx": request.form.get('mascella_dx', 0),
            "mascella_sx": request.form.get('mascella_sx', 0),
            "denti": request.form.get('denti', 0),
            "collo": request.form.get('collo', 0),
            "spalla_dx": request.form.get('spalla_dx', 0),
            "spalla_sx": request.form.get('spalla_sx', 0),
            "braccio_superiore_dx": request.form.get('braccio_superiore_dx', 0),
            "braccio_superiore_sx": request.form.get('braccio_superiore_sx', 0),
            "braccio_inferiore_dx": request.form.get('braccio_inferiore_dx', 0),
            "braccio_inferiore_sx": request.form.get('braccio_inferiore_sx', 0),
            "gomito_dx": request.form.get('gomito_dx', 0),
            "gomito_sx": request.form.get('gomito_sx', 0),
            "polso_dx": request.form.get('polso_dx', 0),
            "polso_sx": request.form.get('polso_sx', 0),
            "mano_dx": request.form.get('mano_dx', 0),
            "mano_sx": request.form.get('mano_sx', 0),
            "dito_dx": request.form.get('dito_dx', 0),
            "dito_sx": request.form.get('dito_sx', 0),
            "petto_superiore_dx": request.form.get('petto_superiore_dx', 0),
            "petto_superiore_sx": request.form.get('petto_superiore_sx', 0),
            "petto_centrale": request.form.get('petto_centrale', 0),
            "pancia": request.form.get('pancia', 0),
            "addome_superiore": request.form.get('addome_superiore', 0),
            "addome_inferiore": request.form.get('addome_inferiore', 0),
            "schiena_superiore_dx": request.form.get('schiena_superiore_dx', 0),
            "schiena_superiore_sx": request.form.get('schiena_superiore_sx', 0),
            "schiena_inferiore_dx": request.form.get('schiena_inferiore_dx', 0),
            "schiena_inferiore_sx": request.form.get('schiena_inferiore_sx', 0),
            "sedere_dx": request.form.get('sedere_dx', 0),
            "sedere_sx": request.form.get('sedere_sx', 0),
            "gamba_superiore_dx": request.form.get('gamba_superiore_dx', 0),
            "gamba_superiore_sx": request.form.get('gamba_superiore_sx', 0),
            "ginocchio_dx": request.form.get('ginocchio_dx', 0),
            "ginocchio_sx": request.form.get('ginocchio_sx', 0),
            "polpaccio_dx": request.form.get('polpaccio_dx', 0),
            "polpaccio_sx": request.form.get('polpaccio_sx', 0),
            "caviglia_dx": request.form.get('caviglia_dx', 0),
            "caviglia_sx": request.form.get('caviglia_sx', 0),
            "piede_dx": request.form.get('piede_dx', 0),
            "piede_sx": request.form.get('piede_sx', 0)
        }

        query = f"""
            INSERT INTO Dolore (
                data, ora, viso, testa_fronte, testa_tempie_dx, testa_tempie_sx, testa_sommità,
                testa_occipite, nuca, mascella_dx, mascella_sx, denti, collo, spalla_dx, spalla_sx,
                braccio_superiore_dx, braccio_superiore_sx, braccio_inferiore_dx, braccio_inferiore_sx,
                gomito_dx, gomito_sx, polso_dx, polso_sx, mano_dx, mano_sx, dito_dx, dito_sx,
                petto_superiore_dx, petto_superiore_sx, petto_centrale, pancia, addome_superiore,
                addome_inferiore, schiena_superiore_dx, schiena_superiore_sx, schiena_inferiore_dx,
                schiena_inferiore_sx, sedere_dx, sedere_sx, gamba_superiore_dx, gamba_superiore_sx,
                ginocchio_dx, ginocchio_sx, polpaccio_dx, polpaccio_sx, caviglia_dx, caviglia_sx,
                piede_dx, piede_sx
            ) VALUES (
                %(data)s, %(ora)s, %(viso)s, %(testa_fronte)s, %(testa_tempie_dx)s, %(testa_tempie_sx)s, %(testa_sommità)s,
                %(testa_occipite)s, %(nuca)s, %(mascella_dx)s, %(mascella_sx)s, %(denti)s, %(collo)s, %(spalla_dx)s, %(spalla_sx)s,
                %(braccio_superiore_dx)s, %(braccio_superiore_sx)s, %(braccio_inferiore_dx)s, %(braccio_inferiore_sx)s,
                %(gomito_dx)s, %(gomito_sx)s, %(polso_dx)s, %(polso_sx)s, %(mano_dx)s, %(mano_sx)s, %(dito_dx)s, %(dito_sx)s,
                %(petto_superiore_dx)s, %(petto_superiore_sx)s, %(petto_centrale)s, %(pancia)s, %(addome_superiore)s,
                %(addome_inferiore)s, %(schiena_superiore_dx)s, %(schiena_superiore_sx)s, %(schiena_inferiore_dx)s,
                %(schiena_inferiore_sx)s, %(sedere_dx)s, %(sedere_sx)s, %(gamba_superiore_dx)s, %(gamba_superiore_sx)s,
                %(ginocchio_dx)s, %(ginocchio_sx)s, %(polpaccio_dx)s, %(polpaccio_sx)s, %(caviglia_dx)s, %(caviglia_sx)s,
                %(piede_dx)s, %(piede_sx)s
            )
        """

        try:
            with conn.cursor() as cursor:
                cursor.execute(query, dati)
            conn.commit()
            return jsonify({"success": True, "message": "Inserimento avvenuto con successo!"})
        except Exception as e:
            return jsonify({"success": False, "message": f"Errore: {str(e)}"})

    return render_template('inserisci_dolore.html')

@app.route("/inserisci_umore", methods=["GET", "POST"])
@login_required
def inserisci_umore():
    if request.method == "POST":
        data = request.form.get("data", 0)
        ora = request.form.get("ora", 0)
        ansia = request.form.get("ansia", 0)
        energia = request.form.get("energia", 0)
        soddisfazione = request.form.get("soddisfazione", 0)
        felicita = request.form.get("felicita", 0)
        stress = request.form.get("stress", 0)

        try:
            cursor.execute("""
                INSERT INTO Umore (data, ora, ansia, energia, soddisfazione, felicita, stress)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (data, ora, ansia, energia, soddisfazione, felicita, stress))
            conn.commit()
            return jsonify({"success": True, "message": "Inserimento umore avvenuto con successo!"})
        except Exception as e:
            return jsonify({"success": False, "message": f"Errore: {str(e)}"})

    return render_template("inserisci_umore.html")

@app.route("/inserisci_sonno", methods=["GET", "POST"])
@login_required
def inserisci_sonno():
    if request.method == "POST":
        data = request.form.get("data", "")
        ora = request.form.get("ora", "")
        data_inizio = request.form.get("data_inizio_sonno", "")
        ora_inizio = request.form.get("ora_inizio", "")
        data_fine = request.form.get("data_fine_sonno", "")
        ora_fine = request.form.get("ora_fine", "")
        qualita = request.form.get("qualita", 0)

        # DEBUG: stampa i dati ricevuti
        print("---- DEBUG: Dati ricevuti ----")
        print(f"Data inserimento: {data}")
        print(f"Ora inserimento: {ora}")
        print(f"Data inizio sonno: {data_inizio}")
        print(f"Ora inizio sonno: {ora_inizio}")
        print(f"Data fine sonno: {data_fine}")
        print(f"Ora fine sonno: {ora_fine}")
        print(f"Qualità sonno: {qualita}")

        try:
            if len(ora_inizio) == 5:
                ora_inizio += ":00"
            if len(ora_fine) == 5:
                ora_fine += ":00"
            if len(ora) == 5:
                ora += ":00"

            datetime_inizio = f"{data_inizio} {ora_inizio}"
            datetime_fine = f"{data_fine} {ora_fine}"

            print(f"Datetime inizio: {datetime_inizio}")
            print(f"Datetime fine: {datetime_fine}")

            ore_di_sonno = calcola_ore_sonno(datetime_inizio, datetime_fine)
            print(f"Ore di sonno calcolate: {ore_di_sonno}")

            cursor.execute("""
                INSERT INTO Sonno (data, ora, data_inizio, ora_inizio, data_fine, ora_fine, qualita, OreDiSonno)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (data, ora, data_inizio, ora_inizio, data_fine, ora_fine, qualita, ore_di_sonno))
            
            conn.commit()
            print("Inserimento avvenuto con successo.")

            return jsonify({"success": True, "message": "Inserimento sonno avvenuto con successo!"})

        except Exception as e:
            print(f"Errore durante l'inserimento: {str(e)}")  # DEBUG errore
            return jsonify({"success": False, "message": f"Errore: {str(e)}"})

    return render_template("inserisci_sonno.html")



@app.route("/inserisci_nota", methods=["GET", "POST"])
@login_required
def inserisci_nota():
    if request.method == "POST":
        data = request.form.get("data", 0)
        ora = request.form.get("ora", 0)
        testo = request.form.get("commento", "")

        try:
            cursor.execute("""
                INSERT INTO Nota (data, ora, testo)
                VALUES (%s, %s, %s)
            """, (data, ora, testo))
            conn.commit()
            return jsonify({"success": True, "message": "Nota inserita con successo!"})
        except Exception as e:
            return jsonify({"success": False, "message": f"Errore durante l'inserimento: {str(e)}"})

    return render_template("inserisci_nota.html")

# --- VISUALIZZAZIONE ---
@app.route("/visualizza_dolore")
@login_required
def visualizza_dolore():
    cursor.execute("SELECT * FROM Dolore")
    dati = cursor.fetchall()
    return render_template("visualizza_dolore.html", dati=dati)

@app.route("/visualizza_umore")
@login_required
def visualizza_umore():
    cursor.execute("SELECT * FROM Umore")
    dati = cursor.fetchall()
    return render_template("visualizza_umore.html", dati=dati)

@app.route("/visualizza_sonno")
@login_required
def visualizza_sonno():
    cursor.execute("SELECT * FROM Sonno")
    dati = cursor.fetchall()
    return render_template("visualizza_sonno.html", dati=dati)

@app.route("/visualizza_nota")
@login_required
def visualizza_nota():
    cursor.execute("SELECT * FROM Nota ORDER BY data DESC, ora DESC")
    dati = cursor.fetchall()
    return render_template("visualizza_nota.html", dati=dati)


# --- GENERAZIONE REPORT EXCEL
@app.route("/scarica_report_excel")
@login_required
def scarica_report_excel():
    due_settimane_fa = datetime.now() - timedelta(weeks=2)
    data_limite = due_settimane_fa.strftime('%Y-%m-%d')
    data_inizio = "2025-06-01"  # replace with actual start date
    data_fine = datetime.now().strftime('%Y-%m-%d')

    cursor.execute("SELECT * FROM Dolore WHERE data >= %s", (data_limite,))
    dolore = cursor.fetchall()
    colonne_dolore = [desc[0] for desc in cursor.description]
    df_dolore = pd.DataFrame(dolore, columns=colonne_dolore)
    dati_dolore = df_dolore.to_dict(orient="records")

    cursor.execute("SELECT * FROM Umore WHERE data >= %s", (data_limite,))
    umore = cursor.fetchall()
    colonne_umore = [desc[0] for desc in cursor.description]
    df_umore = pd.DataFrame(umore, columns=colonne_umore)
    dati_umore = df_umore.to_dict(orient="records")

    cursor.execute("SELECT * FROM Sonno WHERE data >= %s", (data_limite,))
    sonno = cursor.fetchall()
    colonne_sonno = [desc[0] for desc in cursor.description]
    df_sonno = pd.DataFrame(sonno, columns=colonne_sonno)
    dati_sonno = df_sonno.to_dict(orient="records")

    output = genera_report_excel_unico(
        dati_dolore, dati_umore, dati_sonno, data_inizio, data_fine
    )

    filename = f"Report Dario Trinchese ({data_inizio}) ({data_fine}).xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def genera_report_excel_unico(data_dolore, data_umore, data_sonno, data_inizio, data_fine):
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Report Formattato"

    # Strong color shades (0 to 5)
    red_shades = {
        0: "FFFFFF",
        1: "FFB3B3",
        2: "FF6666",
        3: "FF1A1A",
        4: "CC0000",
        5: "800000"
    }
    green_shades = {
        0: "FFFFFF",
        1: "B3FFB3",
        2: "66FF66",
        3: "1AFF1A",
        4: "00CC00",
        5: "006600"
    }
    cyan_shades = {
        0: "FFFFFF",
        1: "B3FFFF",
        2: "66FFFF",
        3: "1AFFFF",
        4: "00CCCC",
        5: "006666"
    }

    # Border styles
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    thick = Side(border_style="thick", color="000000")

    # --- INFO ROW ---
    # We'll need to know the widest table for merging cells:
    def get_final_columns_dolore():
        fixed_columns = ["data", "ora"]
        macro_areas = {
            "Testa": ["viso", "testa_fronte", "testa_tempie_dx", "testa_tempie_sx", "testa_sommità", "testa_occipite", "nuca"],
            "Mascella e Denti": ["mascella_dx", "mascella_sx", "denti"],
            "Collo e Spalle": ["collo", "spalla_dx", "spalla_sx"],
            "Braccia e Mani": [
                "braccio_superiore_dx", "braccio_superiore_sx", "braccio_inferiore_dx", "braccio_inferiore_sx",
                "gomito_dx", "gomito_sx", "polso_dx", "polso_sx", "mano_dx", "mano_sx", "dito_dx", "dito_sx"
            ],
            "Torace e Addome": ["petto_superiore_dx", "petto_superiore_sx", "petto_centrale", "pancia",
                                "addome_superiore", "addome_inferiore"],
            "Schiena": ["schiena_superiore_dx", "schiena_superiore_sx", "schiena_inferiore_dx", "schiena_inferiore_sx",
                        "sedere_dx", "sedere_sx"],
            "Gambe": ["gamba_superiore_dx", "gamba_superiore_sx", "ginocchio_dx", "ginocchio_sx",
                      "polpaccio_dx", "polpaccio_sx", "caviglia_dx", "caviglia_sx", "piede_dx", "piede_sx"]
        }
        final_cols = fixed_columns[:]
        macro_labels = [""] * len(fixed_columns)
        for area, columns in macro_areas.items():
            final_cols.append("")  # spacer
            macro_labels.append("")   # spacer
            final_cols.extend(columns)
            macro_labels.append(area)
            macro_labels.extend([""] * (len(columns) - 1))
        return final_cols

    final_columns_dolore = get_final_columns_dolore()
    final_columns_umore = ["data", "ora", "ansia", "energia", "soddisfazione", "felicita", "stress"]
    final_columns_sonno = ["data", "ora", "ora_inizio", "ora_fine", "qualita"]

    max_columns = max(len(final_columns_dolore), len(final_columns_umore), len(final_columns_sonno))

    info_text = f"Report di Dario Trinchese dal {data_inizio} al {data_fine}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    cell = ws.cell(row=1, column=1, value=info_text)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    current_row = 3  # Start a bit below the info row

    # --- Helper function to apply borders for a table block ---
    def apply_borders(ws, start_row, start_col, n_rows, n_cols):
        for r in range(start_row, start_row + n_rows):
            for c in range(start_col, start_col + n_cols):
                left = thin
                right = thin
                top = thin
                bottom = thin
                if c == start_col:
                    left = medium
                if c == start_col + n_cols - 1:
                    right = medium
                if r == start_row:
                    top = medium
                if r == start_row + n_rows - 1:
                    bottom = medium
                ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    # --- FUNCTION TO WRITE DOLORI TABLE ---
    def write_dolori(start_row):
        ws.cell(row=start_row, column=1, value="Dolori").font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left")
        table_start_row = start_row + 2  # table headers start here

        # Macro-area labels for row 2 and column headers row 3
        fixed_columns = ["data", "ora"]
        macro_areas = {
            "Testa": ["viso", "testa_fronte", "testa_tempie_dx", "testa_tempie_sx", "testa_sommità", "testa_occipite", "nuca"],
            "Mascella e Denti": ["mascella_dx", "mascella_sx", "denti"],
            "Collo e Spalle": ["collo", "spalla_dx", "spalla_sx"],
            "Braccia e Mani": [
                "braccio_superiore_dx", "braccio_superiore_sx", "braccio_inferiore_dx", "braccio_inferiore_sx",
                "gomito_dx", "gomito_sx", "polso_dx", "polso_sx", "mano_dx", "mano_sx", "dito_dx", "dito_sx"
            ],
            "Torace e Addome": ["petto_superiore_dx", "petto_superiore_sx", "petto_centrale", "pancia",
                                "addome_superiore", "addome_inferiore"],
            "Schiena": ["schiena_superiore_dx", "schiena_superiore_sx", "schiena_inferiore_dx", "schiena_inferiore_sx",
                        "sedere_dx", "sedere_sx"],
            "Gambe": ["gamba_superiore_dx", "gamba_superiore_sx", "ginocchio_dx", "ginocchio_sx",
                      "polpaccio_dx", "polpaccio_sx", "caviglia_dx", "caviglia_sx", "piede_dx", "piede_sx"]
        }

        final_columns = fixed_columns[:]
        macro_labels = [""] * len(fixed_columns)
        for area, columns in macro_areas.items():
            final_columns.append("")  # spacer
            macro_labels.append("")   # spacer
            final_columns.extend(columns)
            macro_labels.append(area)
            macro_labels.extend([""] * (len(columns) - 1))

        max_col = len(final_columns)

        # Title row 2 (macro-area headers)
        for col_idx, label in enumerate(macro_labels, start=1):
            if label != "":
                span = 1
                while col_idx + span - 1 < len(macro_labels) and macro_labels[col_idx + span - 1] == "":
                    span += 1
                end_col = col_idx + span - 1
                if end_col > col_idx:
                    ws.merge_cells(start_row=table_start_row, start_column=col_idx, end_row=table_start_row, end_column=end_col)
                cell = ws.cell(row=table_start_row, column=col_idx, value=label)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Skip next merged columns
                for skip_col in range(col_idx + 1, end_col + 1):
                    macro_labels[skip_col - 1] = None
            if macro_labels[col_idx - 1] is None:
                continue

        # Title row 3 (column names)
        col_headers_row = table_start_row + 1
        for col_idx, col_name in enumerate(final_columns, start=1):
            if col_name == "":
                # Just empty spacer column
                ws.cell(row=col_headers_row, column=col_idx, value="")
            else:
                ws.cell(row=col_headers_row, column=col_idx, value=col_name)
            ws.cell(row=col_headers_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=col_headers_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Write data rows
        data_start_row = col_headers_row + 1
        for row_idx, record in enumerate(data_dolore, start=data_start_row):
            for col_idx, col_name in enumerate(final_columns, start=1):
                if col_name == "":
                    ws.cell(row=row_idx, column=col_idx, value=None)
                    continue
                val = record.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=val)

                # Color numeric 0-5 in red shades (only if val is int and in 0..5)
                if isinstance(val, (int, float)) and (0 <= val <= 5) and val == int(val):
                    hex_color = red_shades[int(val)]
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        n_rows = len(data_dolore) + (data_start_row - table_start_row) + 1
        n_cols = max_col
        apply_borders(ws, table_start_row, 1, n_rows, n_cols)

        return table_start_row, n_rows, n_cols

    # --- FUNCTION TO WRITE UMORE TABLE ---
    def write_umore(start_row):
        ws.cell(row=start_row, column=1, value="Umore").font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left")
        table_start_row = start_row + 2

        columns = ["data", "ora", "ansia", "energia", "soddisfazione", "felicita", "stress"]
        max_col = len(columns)

        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=table_start_row, column=col_idx, value=col_name)
            ws.cell(row=table_start_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=table_start_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Data rows
        data_start_row = table_start_row + 1
        for row_idx, record in enumerate(data_umore, start=data_start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                val = record.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=val)

                # Color numeric 0-5 in green shades (only if val is int and in 0..5)
                if isinstance(val, (int, float)) and (0 <= val <= 5) and val == int(val):
                    hex_color = green_shades[int(val)]
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        n_rows = len(data_umore) + (data_start_row - table_start_row) + 1
        n_cols = max_col
        apply_borders(ws, table_start_row, 1, n_rows, n_cols)

        return table_start_row, n_rows, n_cols

    # --- FUNCTION TO WRITE SONNO TABLE ---
    def write_sonno(start_row):
        ws.cell(row=start_row, column=1, value="Sonno").font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left")
        table_start_row = start_row + 2

        columns = ["data", "ora", "ora_inizio", "ora_fine", "qualita"]
        max_col = len(columns)

        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=table_start_row, column=col_idx, value=col_name)
            ws.cell(row=table_start_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=table_start_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Data rows
        data_start_row = table_start_row + 1
        for row_idx, record in enumerate(data_sonno, start=data_start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                val = record.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=val)

                # Color numeric 0-5 in cyan shades (only if val is int and in 0..5)
                if isinstance(val, (int, float)) and (0 <= val <= 5) and val == int(val):
                    hex_color = cyan_shades[int(val)]
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        n_rows = len(data_sonno) + (data_start_row - table_start_row) + 1
        n_cols = max_col
        apply_borders(ws, table_start_row, 1, n_rows, n_cols)

        return table_start_row, n_rows, n_cols

    # Write Dolori table
    start_dolori = current_row
    dolori_header_row, dolori_n_rows, dolori_n_cols = write_dolori(start_dolori)

    # Leave 3 blank rows between tables
    start_umore = dolori_header_row + dolori_n_rows + 3
    umore_header_row, umore_n_rows, umore_n_cols = write_umore(start_umore)

    start_sonno = umore_header_row + umore_n_rows + 3
    sonno_header_row, sonno_n_rows, sonno_n_cols = write_sonno(start_sonno)

    # Adjust column widths to fit contents for all columns used in all tables
    max_cols_used = max(dolori_n_cols, umore_n_cols, sonno_n_cols)
    for col_idx in range(1, max_cols_used + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- MAIN ---
if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)
